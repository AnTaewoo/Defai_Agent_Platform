"""[P1] Excel(.xlsx) 파서. openpyxl로 직접 제어 — LangChain 로더 위임 금지."""

from __future__ import annotations

from openpyxl import load_workbook

from ..types import ParsedDocument


def parse_excel(path: str) -> ParsedDocument:
    """xlsx 파일을 ParsedDocument로 변환.

    시트마다 표(tables)로 보존하고, 시트명을 text_blocks에 남겨 맥락을 유지한다.
    빈 행/빈 셀은 ""로 채워 표 구조(행렬 형태)를 보존한다.
    """
    wb = load_workbook(path, data_only=True, read_only=True)

    text_blocks: list[str] = []
    tables: list[list[list[str]]] = []

    for sheet in wb.worksheets:
        text_blocks.append(f"# {sheet.title}")

        rows: list[list[str]] = []
        for row in sheet.iter_rows():
            cells = ["" if cell.value is None else str(cell.value) for cell in row]
            if any(c != "" for c in cells):
                rows.append(cells)

        if rows:
            tables.append(rows)

    return ParsedDocument(
        source=path,
        doc_type="xlsx",
        text_blocks=text_blocks,
        tables=tables,
    )
